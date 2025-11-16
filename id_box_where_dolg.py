import csv
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# === НАСТРОЙКИ ОФОРМЛЕНИЯ ===
COLUMN_WIDTHS = [15, 25, 30, 20]
FONT_NAME = "Calibri"
FONT_SIZE = 11
HEADER_FONT_SIZE = 12
BORDER_STYLE = Side(border_style="thin", color="000000")

# === ДАННЫЕ ===
file_path = 'example.csv'
target_column_index = 17
desired_values = ["Должник", "Должен за 2025", "Должен за 2024-2025"]

values_to_collect = []

# Функция для безопасного преобразования в float
def safe_float(value):
    if not value:
        return 0.0
    value = value.strip().replace(',', '.')
    try:
        return float(value)
    except ValueError:
        return 0.0

with open(file_path, 'r', encoding="utf-8") as file:
    reader = csv.reader(file, delimiter=';')
    for row in reader:
        if len(row) > target_column_index:
            if row[target_column_index] in desired_values:
                # Проверяем дубли по номеру гаража
                if not any(r[0] == row[0] for r in values_to_collect):
                    # Обработка ФИО
                    st = ' '.join(row[1].split()[1:])
                    if '.' not in st:
                        parts = st.split()
                        if len(parts) >= 2:
                            st = ' '.join(
                                [parts[0]] + [f"{name[0]}." for name in parts[1:-1]] + [f" {parts[-1][0]}."]
                            )

                    # Обновляем статус
                    status = row[target_column_index]
                    if status == "Должник":
                        status = "Должен до 2024 и за 2024-2025"

                    # Преобразуем долг в число
                    debt_raw = row[target_column_index - 4] if len(row) > target_column_index - 4 else ""
                    debt_float = safe_float(debt_raw)

                    values_to_collect.append([row[0], st, status, debt_float])

# Убираем заголовок, если он попал
if values_to_collect and len(values_to_collect[0]) == 4 and values_to_collect[0][0].lower() == 'номер':
    values_to_collect.pop(0)

# Сортировка по статусу
sorted_list = sorted(values_to_collect, key=lambda x: x[2])

# === ВЫВОД В КОНСОЛЬ ===
print("\nЗначения из первого столбца, соответствующие условию:")
print("-" * 80)
for item in sorted_list:
    # Форматируем долг как число с 2 знаками
    debt_str = f"{item[3]:,.2f}".replace(",", " ") if item[3] != 0 else "0.00"
    print(f"{item[0]} | {item[1]} | {item[2]} | {debt_str}")
print("-" * 80)

# === СОЗДАНИЕ EXCEL ===
workbook = openpyxl.Workbook()
font_default = Font(name=FONT_NAME, size=FONT_SIZE)
font_header = Font(name=FONT_NAME, size=HEADER_FONT_SIZE, bold=True)
align_center = Alignment(horizontal="center", vertical="center")
align_left = Alignment(horizontal="left", vertical="center")
border = Border(left=BORDER_STYLE, right=BORDER_STYLE, top=BORDER_STYLE, bottom=BORDER_STYLE)

headers = ['Номер гаража', 'ФИО', 'Статус должника', 'Остаток долга']

def format_sheet(sheet, data_list, title_modifiers=None):
    sheet.title = title_modifiers.get('title', sheet.title) if title_modifiers else sheet.title
    sheet.append(headers)

    # Заголовки
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.font = font_header
        cell.alignment = align_center
        cell.border = border

    # Данные
    for row_idx, data in enumerate(data_list, start=2):
        for col_idx, value in enumerate(data, 1):
            cell = sheet.cell(row=row_idx, column=col_idx, value=value)
            cell.font = font_default
            cell.border = border

            # Выравнивание
            if col_idx == 1 or col_idx == 2:
                cell.alignment = align_left
            else:
                cell.alignment = align_center

            # Форматирование числа (долг)
            if col_idx == 4 and isinstance(value, (int, float)):
                cell.number_format = '#,##0.00'

            # Изменение статуса на втором листе
            if title_modifiers and title_modifiers.get('modify_status') and col_idx == 3:
                cell.value = title_modifiers['modify_status']

    # Ширина столбцов
    for i, width in enumerate(COLUMN_WIDTHS, 1):
        sheet.column_dimensions[get_column_letter(i)].width = width

    # Высота строк
    for row in range(2, sheet.max_row + 1):
        sheet.row_dimensions[row].height = 18

# === Листы ===
# 1. Все данные
sheet1 = workbook.active
sheet1.title = 'All Data'
format_sheet(sheet1, sorted_list)

# 2. Должники за 2025
sheet2 = workbook.create_sheet('Должники за 2025')
data_2025 = [d.copy() for d in sorted_list if d[2] == 'Должен за 2025']
for d in data_2025:
    d[2] = 'Еще не оплатил за 2025'
format_sheet(sheet2, data_2025, title_modifiers={'modify_status': 'Еще не оплатил за 2025'})

# 3. Должники за 2024-2025
sheet3 = workbook.create_sheet('Должники за 2024-2025')
data_2024_2025 = [d for d in sorted_list if d[2] == 'Должен за 2024-2025']
format_sheet(sheet3, data_2024_2025)

# 4. Должники до 2024
sheet4 = workbook.create_sheet('Должники до 2024')
data_before = [d for d in sorted_list if d[2] not in ['Должен за 2024-2025', 'Должен за 2025']]
format_sheet(sheet4, data_before)

# === ДОБАВЛЕНИЕ ЛИСТА "Сводка" С ТАБЛИЦЕЙ И КРУГОВОЙ ДИАГРАММОЙ ===

from collections import Counter
from openpyxl.chart import PieChart, Reference
import openpyxl.drawing.fill

# === ЗАДАЁМ ВРУЧНУЮ ===
TOTAL_GARAGES = 413  # ← ТУТ МЕНЯЙ НА НУЖНОЕ ЧИСЛО

# Подсчёт должников по статусам
status_counter = Counter()
for item in sorted_list:
    status = item[2]
    if status == "Должен за 2025":
        status_counter["Еще не оплатил за 2025"] += 1
    elif status in ["Должен за 2024-2025", "Должен до 2024 и за 2024-2025"]:
        status_counter["Должен до 2025"] += 1

# Оплачено = Всего - Должников
debtors = len(sorted_list)
paid = TOTAL_GARAGES - debtors

# Заполняем счётчик
status_counter["Еще не оплатил за 2025"] = status_counter.get("Еще не оплатил за 2025", 0)
status_counter["Должен до 2025"] = status_counter.get("Должен до 2025", 0)
status_counter["Все оплатил"] = paid
status_counter["Всего гаражей"] = TOTAL_GARAGES

# === Создание листа "Сводка" ===
summary_sheet = workbook.create_sheet('Сводка', 0)  # На первое место

# Заголовки
summary_sheet['A1'] = 'Статус'
summary_sheet['B1'] = 'Количество'
for cell in ['A1', 'B1']:
    summary_sheet[cell].font = font_header
    summary_sheet[cell].alignment = align_center
    summary_sheet[cell].border = border

# Данные
summary_data = [
    ["Должен до 2025", status_counter["Должен до 2025"]],
    ["Еще не оплатил за 2025", status_counter["Еще не оплатил за 2025"]],
    ["Все оплатил", status_counter["Все оплатил"]],
    ["Всего гаражей", TOTAL_GARAGES]
]

for row_idx, (status, count) in enumerate(summary_data, start=2):
    cell_a = summary_sheet.cell(row=row_idx, column=1, value=status)
    cell_b = summary_sheet.cell(row=row_idx, column=2, value=count)
    cell_a.font = font_default
    cell_b.font = font_default
    cell_a.alignment = align_left
    cell_b.alignment = align_center
    cell_a.border = border
    cell_b.border = border

# === ОФОРМЛЕНИЕ ===
summary_sheet.column_dimensions['A'].width = 28
summary_sheet.column_dimensions['B'].width = 15
for i in range(1, 6):
    summary_sheet.row_dimensions[i].height = 22

print(f"Сводка: Всего гаражей = {TOTAL_GARAGES} | Должников = {debtors} | Оплачено = {paid}")

# === СОХРАНЕНИЕ ===
import datetime
timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
filename = f'dolgniki_{timestamp}.xlsx'

try:
    workbook.save(filename)
    print(f"\nФайл сохранён: {filename}")
    print(f"Всего гаражей: {TOTAL_GARAGES} | Должников: {debtors} | Оплачено: {paid}")
except Exception as e:
    print(f"Ошибка: {e}")